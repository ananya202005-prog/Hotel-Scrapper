"""
Concurrent Hotel Scraper — MakeMyTrip
Uses Scrapling's StealthyFetcher with 2 parallel Chrome instances via
ThreadPoolExecutor (Windows-compatible) to scrape hotel data from 10
MMT hotel detail pages.
"""

import logging
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Optional

import yaml
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  [%(threadName)s]  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("scraper.log", mode="w", encoding="utf-8"),
    ],
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------
@dataclass
class HotelResult:
    url: str
    hotel_name: Optional[str] = None
    price: Optional[str] = None
    rating: Optional[str] = None
    reviews: Optional[str] = None
    error: Optional[str] = None

# ---------------------------------------------------------------------------
# Configuration helpers
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent


def load_urls(yaml_path: str = "urls.yaml") -> list:
    """Load target URLs from the YAML configuration file."""
    path = BASE_DIR / yaml_path
    if not path.exists():
        raise FileNotFoundError(f"urls.yaml not found at {path}")
    with open(path, encoding="utf-8") as f:
        data = yaml.safe_load(f)
    urls = data.get("urls", [])
    if not urls:
        raise ValueError("No URLs found in urls.yaml")
    logger.info("Loaded %d URLs from %s", len(urls), yaml_path)
    return urls


def build_proxy(worker_id: int) -> Optional[dict]:
    """
    Build proxy configuration from environment variables.
    Per-worker vars (PROXY_SERVER_1, PROXY_SERVER_2) take priority
    over shared vars (PROXY_SERVER). Returns None when no proxy configured.
    """
    server   = os.getenv(f"PROXY_SERVER_{worker_id}")   or os.getenv("PROXY_SERVER")
    username = os.getenv(f"PROXY_USERNAME_{worker_id}") or os.getenv("PROXY_USERNAME")
    password = os.getenv(f"PROXY_PASSWORD_{worker_id}") or os.getenv("PROXY_PASSWORD")

    if not server:
        return None

    proxy: dict = {"server": server}
    if username:
        proxy["username"] = username
    if password:
        proxy["password"] = password
    return proxy


# ---------------------------------------------------------------------------
# Selectors — MakeMyTrip hotel detail page
# ---------------------------------------------------------------------------
HOTEL_NAME_SELECTORS = [
    "h1[data-testid='hotel-name']",
    "h1.hotelName",
    "h1",
    "[class*='hotelName']",
    "[class*='hotel-name']",
    "[class*='propertyName']",
    ".hotel_name",
]

PRICE_SELECTORS = [
    "[data-testid='display-price']",
    "[class*='priceSection'] [class*='actualPrice']",
    "[class*='actual-price']",
    "[class*='displayPrice']",
    "p.actual-price",
    "[class*='price'] strong",
    "[class*='roomPrice']",
]

RATING_SELECTORS = [
    "[data-testid='hotel-rating']",
    "[class*='rating'] [class*='value']",
    "[class*='ratingValue']",
    "[class*='hotelRating']",
    ".rating-value",
    "span[class*='rating']",
]

REVIEW_SELECTORS = [
    "[data-testid='review-count']",
    "[class*='reviewCount']",
    "[class*='review-count']",
    "span[class*='review']",
    "[class*='ratingCount']",
]


def _first_text(page, selectors: list) -> Optional[str]:
    """Try each selector; return stripped text of the first match."""
    for selector in selectors:
        try:
            el = page.css_first(selector)
            if el and el.text:
                text = el.text.strip()
                if text:
                    return text
        except Exception:
            continue
    return None


# ---------------------------------------------------------------------------
# Worker — runs in a thread
# ---------------------------------------------------------------------------
MAX_RETRIES   = 3
RETRY_DELAY   = 5        # seconds
FETCH_TIMEOUT = 60_000   # milliseconds


def scrape_batch(worker_id: int, urls: list) -> list:
    """
    Scrape a batch of URLs in one thread, each with its own browser session.
    Returns a list of dicts (one per URL).
    """
    log = logging.getLogger(f"worker.{worker_id}")

    # Worker-level log file
    fh = logging.FileHandler(f"worker_{worker_id}.log", mode="w", encoding="utf-8")
    fh.setFormatter(logging.Formatter(
        "%(asctime)s  %(levelname)-8s  %(message)s", datefmt="%H:%M:%S"
    ))
    log.addHandler(fh)

    try:
        from scrapling.fetchers.stealth_chrome import StealthyFetcher
    except Exception as exc:
        log.error("Failed to import StealthyFetcher: %s", exc)
        return [asdict(HotelResult(url=u, error=f"Import error: {exc}")) for u in urls]

    proxy = build_proxy(worker_id)
    if proxy:
        log.info("Worker %d using proxy: %s", worker_id, proxy.get("server"))
    else:
        log.info("Worker %d running WITHOUT proxy", worker_id)

    results = []

    for url in urls:
        log.info("Worker %d → %s", worker_id, url)
        result = HotelResult(url=url)

        for attempt in range(1, MAX_RETRIES + 1):
            try:
                log.info("  Attempt %d/%d — launching browser …", attempt, MAX_RETRIES)
                page = StealthyFetcher.fetch(
                    url,
                    headless=True,
                    disable_resources=False,   # keep resources — MMT needs them to render
                    block_ads=True,
                    hide_canvas=True,
                    block_webrtc=True,
                    network_idle=True,
                    load_dom=True,
                    timeout=FETCH_TIMEOUT,
                    wait=3000,                 # extra 3s after load for JS to settle
                    proxy=proxy,
                    google_search=True,
                    locale="en-IN",
                    timezone_id="Asia/Kolkata",
                    extra_flags=[
                        "--disable-http2",
                        "--no-sandbox",
                        "--disable-dev-shm-usage",
                    ],
                )

                result.hotel_name = _first_text(page, HOTEL_NAME_SELECTORS)
                result.price      = _first_text(page, PRICE_SELECTORS)
                result.rating     = _first_text(page, RATING_SELECTORS)
                result.reviews    = _first_text(page, REVIEW_SELECTORS)
                result.error      = None

                log.info(
                    "Worker %d OK | name=%r  price=%r  rating=%r  reviews=%r",
                    worker_id, result.hotel_name, result.price,
                    result.rating, result.reviews,
                )
                break  # success

            except Exception as exc:
                log.warning(
                    "Worker %d attempt %d/%d FAILED: %s",
                    worker_id, attempt, MAX_RETRIES, exc,
                )
                if attempt < MAX_RETRIES:
                    time.sleep(RETRY_DELAY * attempt)
                else:
                    result.error = str(exc)
                    log.error("Worker %d gave up on %s", worker_id, url)

        results.append(asdict(result))

    log.info("Worker %d finished — %d URLs processed", worker_id, len(results))
    return results


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------
OUTPUT_FIELDS = ["url", "hotel_name", "price", "rating", "reviews", "error"]


def save_xlsx(results: list, path: str = "output.xlsx") -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hotels"
    ws.append(OUTPUT_FIELDS)

    # Style header
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in results:
        ws.append([row.get(f) for f in OUTPUT_FIELDS])

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

    wb.save(path)
    logger.info("Excel saved → %s", path)


# ---------------------------------------------------------------------------
# Main orchestrator
# ---------------------------------------------------------------------------
NUM_WORKERS = 2


def distribute(urls: list, n: int) -> list:
    """Round-robin split across n buckets."""
    buckets = [[] for _ in range(n)]
    for i, url in enumerate(urls):
        buckets[i % n].append(url)
    return buckets


def main() -> None:
    load_dotenv()
    start = time.time()

    logger.info("=" * 60)
    logger.info("Concurrent Hotel Scraper — starting")
    logger.info("Workers : %d  (threads)", NUM_WORKERS)
    logger.info("=" * 60)

    # Verify Scrapling import early so errors are visible
    try:
        from scrapling.fetchers.stealth_chrome import StealthyFetcher  # noqa: F401
        logger.info("Scrapling import OK")
    except Exception as exc:
        logger.error("Cannot import StealthyFetcher: %s", exc)
        logger.error("Run:  python -m patchright install chromium")
        sys.exit(1)

    urls    = load_urls()
    batches = distribute(urls, NUM_WORKERS)

    for i, batch in enumerate(batches, 1):
        logger.info("Worker %d will scrape %d URLs", i, len(batch))

    all_results = []

    with ThreadPoolExecutor(max_workers=NUM_WORKERS, thread_name_prefix="Worker") as pool:
        futures = {
            pool.submit(scrape_batch, worker_id, batch): worker_id
            for worker_id, batch in enumerate(batches, start=1)
        }
        logger.info("Both browser threads started — waiting for results …")

        for future in as_completed(futures):
            worker_id = futures[future]
            try:
                batch_results = future.result()
                all_results.extend(batch_results)
                logger.info(
                    "Worker %d completed — returned %d results",
                    worker_id, len(batch_results),
                )
            except Exception as exc:
                logger.error(
                    "Worker %d raised an unhandled exception: %s",
                    worker_id, exc, exc_info=True,
                )

    # Restore original URL order
    url_order = {url: idx for idx, url in enumerate(urls)}
    all_results.sort(key=lambda r: url_order.get(r["url"], 9999))

    save_xlsx(all_results)

    elapsed = time.time() - start
    success = sum(1 for r in all_results if not r.get("error"))
    failed  = len(all_results) - success

    logger.info("=" * 60)
    logger.info("Execution summary")
    logger.info("  Total URLs : %d", len(urls))
    logger.info("  Succeeded  : %d", success)
    logger.info("  Failed     : %d", failed)
    logger.info("  Duration   : %.1f s", elapsed)
    logger.info("=" * 60)

    if failed:
        logger.warning("Failed URLs:")
        for r in all_results:
            if r.get("error"):
                logger.warning("  %s\n    → %s", r["url"], r["error"])


if __name__ == "__main__":
    main()